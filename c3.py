#!/usr/bin/env python3
# -*- coding: utf-8 -*-
from openpyxl import Workbook
from openpyxl import load_workbook

wb = load_workbook('pandas2.xlsx')
ws = wb.active
ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=4)
wb.save('pandas2.xlsx')
