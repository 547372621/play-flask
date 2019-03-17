#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

df = pd.read_csv('aaa.txt')
p = df.pivot_table(index=['time'],columns=['dt'],values=['value1','value2','value3'])\
    .reorder_levels([1,0], axis=1)

# 获取全部列数
column_count = len(p.columns.labels[0])
# 获取二级分类的个数
second_index_count = len(p.columns.levels[1])
# 生成新的索引
column_indexes = sorted(range(column_count),key=lambda x:x%second_index_count)
df = p.iloc[:,column_indexes]
print(df)

wb = Workbook()
ws = wb.active

for r in dataframe_to_rows(df, index=True, header=True):
    ws.append(r)

for r in dataframe_to_rows(df, index=True, header=True):
    ws.append(r)

# for cell in ws['A'] + ws[1]:
#     cell.style = 'Pandas'

wb.save("pandas2.xlsx")